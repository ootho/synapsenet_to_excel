from bs4 import BeautifulSoup
from files_io import duplicates_check, kpp_check
import requests
import json
import xlsxwriter
import openpyxl
from openpyxl.styles import PatternFill

import bs4_sandbox

# получаем html по списку ИНН
def bs4_scraper(path : str):
    with open(path,'r') as file:
        data = file.read()
        tax_id_list = json.loads(data)
    
    d = 'https://synapsenet.ru'
    for i in tax_id_list:
        url = f'{d}/searchorganization/proverka-kontragentov?query={i}'
        page = requests.get(url)

    soup = BeautifulSoup(page.content, 'html.parser')

    # Проверить больше ли 1 ИНН мы обнаружили
    check_if_many = soup.find_all(class_="org-pcl-go-but")
    if check_if_many!=None:
        hrefs = [i['href'] for i in soup.find_all(class_="org-pcl-go-but")]

        url = f'{d}/searchorganization/proverka-kontragentov?query={hrefs[0]}'
        page = requests.get(url)
        soup = BeautifulSoup(page.content, 'html.parser')

    # Проверить нашли ли мы что-нибудь
    # if soup.contents.find("открыть карточку") != None:
        # print('Найдено более 1 организации с искомым ИНН')
        # return

    with open('html.html','w') as file:
        file.write(str(soup.contents))
    return

# bs4_scraper('file.txt')

def write_xlsx(data):
    try:
        workbook = openpyxl.load_workbook('final/report.xlsx')
    except Exception as ex:
        print(ex)
        workbook = openpyxl.Workbook(write_only=False)
    
    worksheet = workbook.active

    # Если лист пустой добавляем заголовки
    row = worksheet.max_row
    if row == 1:
        for idx, column in enumerate(data.keys()):
            worksheet.cell(row=row, column=idx+1, value=column)

    # Добавляем данные в свободную строку
    row += 1
    for idx, value in enumerate(data.values()):
        worksheet.cell(row=row, column=idx+1, value=value)

    #/////////////////////////////////////////////////////////////////////////////
    '''
    если есть, то добавляем новые данные
    ели данные пустые, то меняем цвет ячеек
    '''
    #/////////////////////////////////////////////////////////////////////////////
    # Iterate over the data and write it out row by row.

    # Сохраняем книгу
    workbook.save('final/report.xlsx')
    return

def xlsx_fill_color(data):
    try:
        workbook = openpyxl.load_workbook('final/report.xlsx')
    except Exception as ex:
        print(ex)
        workbook = openpyxl.Workbook(write_only=False)
    
    worksheet = workbook.active

    # Если лист пустой добавляем заголовки
    row = worksheet.max_row
    if row == 1:
        for idx, column in enumerate(data.keys()):
            worksheet.cell(row=row, column=idx+1, value=column)

    # Добавляем данные в свободную строку
    row += 1
    for idx, value in enumerate(data.values()):
        cell = worksheet.cell(row=row, column=idx+1, value=value)
        # if not value:
        # Создаем объект Fill с красным цветом
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        # Присваиваем Fill ячейке
        cell.fill = red_fill
    #/////////////////////////////////////////////////////////////////////////////
    '''
    если есть, то добавляем новые данные
    ели данные пустые, то меняем цвет ячеек
    '''
    #/////////////////////////////////////////////////////////////////////////////
    # Iterate over the data and write it out row by row.

    # Сохраняем книгу
    workbook.save('final/report.xlsx')
    return

def scraper_main_loop(path):
    with open(path,'r') as file:
        data = file.read()
        tax_id_list = json.loads(data)
    
    # Проходим по списку ИНН
    for tax_id in tax_id_list:

        # Проверяем есть ли данные по этому ИНН в файле Эксель
        if not duplicates_check('final/report.xlsx', 'ИНН', tax_id):

            # Запрашиваем и получаем HTML документ
            url = f'https://synapsenet.ru/searchorganization/proverka-kontragentov?query={tax_id}'

            #  Получаем страницу по ссылке
            page = requests.get(url)
            # Если получили страницу, то парсим
#///////////////////////////////////////////////////////////////////////////////
            soup = BeautifulSoup(page.content, 'html.parser')
            # Если не получили, то всталяем пустую строку и выделяем цветом
            if len(soup.find_all(class_='orgs-not-found'))>0:
                xlsx_fill_color({"ИНН" : tax_id, "КПП" : "123aaa"})
            else:

                # Парсим документ и получаем словарь с данными
                # Проверить больше ли одной органищации по ИНН мы обнаружили
                soups = []
                check_if_many = soup.find_all(class_="org-pcl-go-but")
                if len(check_if_many)>0:
                    # Если мы нашли болье 1 организации по ИНН, то сохраняем сслыки на все их страницы
                    hrefs = [i['href'] for i in soup.find_all(class_="org-pcl-go-but")]

                    d = 'https://synapsenet.ru'
                    for ref in hrefs:
                        url = f'{d}/searchorganization/proverka-kontragentov?query={ref}'
                        page = requests.get(url)
                        soup = BeautifulSoup(page.content, 'html.parser')

                        # Добавляет в список
                        soups.append(soup)
                else:
                    soups.append(soup)
                
                # Записываем данные в файлы (совсем не обязательная процедура)
                # with open('html.html','w') as file:
                #     file.write(str(soup.contents))

                for one_soup in soups:
                    # Собираем необходимые данные с помощью bs4 и получаем словарь
                    data_dict = bs4_sandbox.html_scraper(one_soup)
                    
                    duplicates = False
                    # Если несколько организаций с одним ИНН проверяем по КПП:
                    if len(check_if_many)>0:
                        duplicates = duplicates_check('final/report.xlsx', 'КПП', data_dict['КПП'])

                    # Добавляем значения словаря в файл Эксель
                    if not duplicates: write_xlsx(data_dict)

    return


# "найдено 2 лица"

# scraper_main_loop('file.txt')
