import os
import shutil
import time

import openpyxl
from openpyxl.styles import PatternFill

BASE_DIR = os.getcwd()


def write_xlsm(data):
    # Пробуем открыть файл, если нет, то создаём новый
    try:
        workbook = openpyxl.load_workbook(filename='wip/report.xlsm', 
                                          read_only=False, keep_vba=True)
    except Exception as ex:
        print(ex)
        workbook = openpyxl.Workbook(write_only=False)

    worksheet = workbook.active

    # Если лист пустой, то добавляем заголовки
    row = worksheet.max_row
    if row == 1:
        for idx, column in enumerate(data.keys()):
            worksheet.cell(row=row, column=idx+1, value=column)

    # Если в словаре передаются данные, которым не соответствует ни одна
    # колонка, создаём новую колонку
    for idx, column in enumerate(data.keys()):
        if column not in [cell.value for cell in worksheet[1]]:
            worksheet.cell(row=1, column=worksheet.max_column + 1, value=column)

    row += 1
    for idx, column in enumerate(worksheet[1], start=1):
        if column.value in data:
            cell = worksheet.cell(row=row, column=idx,
                                  value=data[column.value])

        # Если получили только ИНН и timestamp, значит данные не 
        # обнаружены, пишем в таблицу то что есть и выделаем строку цветом
        if len([x for x in list(data.values()) if x != 'na']) == 2:
            red_fill = PatternFill(
                start_color='FFFF0000', 
                end_color='FFFF0000', 
                fill_type='solid'
                )
            cell.fill = red_fill

    # Сохраняем книгу и закрываем
    workbook.save('wip/report.xlsm')
    workbook.close()
    return


# Проверка есть ли ИНН в файле xlsm, true - найден дубликат, false - нет
def duplicates_check(path: str, name: str, feature: str) -> bool:
    try:
        wb = openpyxl.load_workbook(filename=path, keep_vba=True)
        sheet = wb.active
        # Получение индекса столбца по нужному заголовку (ИНН или КПП)
        column_num = next(sheet.values).index(name) + 1

        # Получаем букву столбца по индексу столбца
        column_letter = openpyxl.utils.cell.get_column_letter(column_num)
        
        # Ищем данные в ячейках соответствующего столбца
        if feature in [cell.value for cell in sheet[column_letter]]:
            # Признак найден
            wb.close()
            return True
        else:
            # Признак не найден
            wb.close()
            return False
        
    except Exception as x:
        # print (x)
        wb.close()
        return False

# Изменение порядка столбцов перед отправкой пользователю
def preparing_to_upload():
    wb = openpyxl.load_workbook(filename='wip/report.xlsm', 
                                read_only=False, keep_vba=True)
    ws = wb.active

    # Сортируем по алфавиту после 22го столбца
    columns = [cell.value for cell in ws[1][22:]]
    columns_sorted = reversed(sorted(columns))
    for column in columns_sorted:
        ws.insert_cols(23)
        col_upd = [cell.value for cell in ws[1][23:]]
        old_index = col_upd.index(column) + 1 + 23
        col = openpyxl.utils.get_column_letter(old_index)
        end_row = ws.max_row
        cell_range = f"{col}1:{col}{end_row}"
        ws.move_range(cell_range, cols=23-old_index)
        ws.delete_cols(old_index)

    # Выбираем столбец с выручкой за последний доступный год
    # и перемещаем на место второго столбца
    end_row = ws.max_row
    ws.insert_cols(2)
    last_revenue = sorted([cell.value for cell in ws[1]
                          if "Выручка" in str(cell.value)])[-1]
    old_index = [i for i, 
                 cell in enumerate(ws[1]) 
                 if last_revenue in str(cell.value)][0] + 1
    col = openpyxl.utils.get_column_letter(old_index)
    cell_range = f"{col}1:{col}{end_row}"
    ws.move_range(cell_range, cols=2-old_index)
    ws.delete_cols(old_index)

    # Удаляем столбец timestamp
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, 
        min_col=1, max_col=ws.max_column
        ):
        for cell in row:
            if cell.value == 'timestamp':
                ws.delete_cols(cell.column)

    wb.save('sent/report.xlsm')
    wb.close()

# Проверяем существует ли файл и когда последний раз модифицирован
# если файл существует и изменён менее суток назад возвращает False
# иначе возвращаем True
def not_exist_or_modified_more_than_day_ago():
    try:
        now = time.time()
        filepath = f"{BASE_DIR}/wip/report.xlsm"
        mtime = os.path.getmtime(filepath)
        return (now - mtime) > (24 * 60 * 60)
    except Exception as x:
        # print(x)
        return True


def delete_file():
    try:
        filepath = f"{BASE_DIR}/wip/report.xlsm"
        os.remove(filepath)
    except Exception as x:
        # print(x)
        pass


def copy_template():
    src_file = f"{BASE_DIR}/template/template.xlsm"
    dst_dir = f"{BASE_DIR}/wip/report.xlsm"
    shutil.copy(src_file, dst_dir)
